from decimal import Decimal

from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Sum, Q
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse

from django.utils import timezone

from fees.models import (
    AcademicSession,
    Term,
    FeeStructure,
    Payment,
)

from parents.models import Parent
from students.models import Student

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# ============================================================
# STAFF DASHBOARD
# ============================================================

@staff_member_required
def dashboard(request):

    total_students = Student.objects.filter(
        is_active=True
    ).count()

    total_parents = Parent.objects.filter(
        is_active=True
    ).count()

    total_fee_structures = FeeStructure.objects.count()

    active_session = AcademicSession.objects.filter(
        is_active=True
    ).first()

    verified_payments = Payment.objects.filter(
        status="VERIFIED"
    )

    pending_payments = Payment.objects.filter(
        status="PENDING"
    )

    rejected_payments = Payment.objects.filter(
        status="REJECTED"
    )

    total_collected = (
        verified_payments.aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0")
    )

    pending_amount = (
        pending_payments.aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0")
    )

    rejected_amount = (
        rejected_payments.aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0")
    )

    expected_fees = Decimal("0")

    if active_session:

        active_students = Student.objects.filter(
            is_active=True
        )

        for student in active_students:

            fee_query = FeeStructure.objects.filter(
                session=active_session,
                student_class=student.student_class,
                student_type=student.student_type,
            )

            if student.student_class.startswith("SSS"):

                fee_query = fee_query.filter(
                    department=student.department
                )

            else:

                fee_query = fee_query.filter(
                    department__isnull=True
                )

            fee = fee_query.first()

            if fee:
                expected_fees += fee.total_fee

    outstanding = expected_fees - total_collected

    if outstanding < Decimal("0"):
        outstanding = Decimal("0")

    recent_payments = (
        Payment.objects
        .select_related(
            "student",
            "fee_structure",
        )
        .order_by(
            "-transaction_date"
        )[:8]
    )

    context = {
        "total_students": total_students,
        "total_parents": total_parents,
        "total_fee_structures": total_fee_structures,

        "active_session": active_session,

        "total_collected": total_collected,
        "pending_amount": pending_amount,
        "rejected_amount": rejected_amount,

        "pending_count": pending_payments.count(),
        "rejected_count": rejected_payments.count(),

        "expected_fees": expected_fees,
        "outstanding": outstanding,

        "recent_payments": recent_payments,
    }

    return render(
        request,
        "dashboard/dashboard.html",
        context
    )


# ============================================================
# STUDENTS
# ============================================================

@staff_member_required
def student_list(request):

    students = (
        Student.objects
        .all()
        .order_by("-registered_at")
    )

    search = request.GET.get(
        "search",
        ""
    ).strip()

    selected_class = request.GET.get(
        "class",
        ""
    )

    selected_status = request.GET.get(
        "status",
        ""
    )

    if search:

        students = students.filter(
            Q(full_name__icontains=search)
            |
            Q(student_id__icontains=search)
            |
            Q(parent_name__icontains=search)
            |
            Q(parent_phone__icontains=search)
            |
            Q(parent_email__icontains=search)
        )

    if selected_class:

        students = students.filter(
            student_class=selected_class
        )

    if selected_status == "active":

        students = students.filter(
            is_active=True
        )

    elif selected_status == "inactive":

        students = students.filter(
            is_active=False
        )

    context = {
        "students": students,

        "search": search,

        "selected_class": selected_class,

        "selected_status": selected_status,

        "class_choices": Student.CLASS_CHOICES,

        "total_students": Student.objects.count(),

        "active_students": Student.objects.filter(
            is_active=True
        ).count(),

        "inactive_students": Student.objects.filter(
            is_active=False
        ).count(),
    }

    return render(
        request,
        "dashboard/student_management.html",
        context
    )


# ============================================================
# ACTIVATE / DEACTIVATE STUDENT
# ============================================================

@staff_member_required
def toggle_student_status(request, student_id):

    student = get_object_or_404(
        Student,
        id=student_id
    )

    if request.method == "POST":

        student.is_active = not student.is_active

        student.save()

    return redirect(
        "dashboard:students"
    )


# ============================================================
# PARENTS
# ============================================================

@staff_member_required
def parent_list(request):

    parents = (
        Parent.objects
        .select_related("user")
        .prefetch_related("students")
        .order_by("-created_at")
    )

    search = request.GET.get(
        "search",
        ""
    ).strip()

    selected_status = request.GET.get(
        "status",
        ""
    ).strip()

    if search:

        parents = parents.filter(
            Q(full_name__icontains=search)
            |
            Q(phone__icontains=search)
            |
            Q(user__username__icontains=search)
            |
            Q(user__email__icontains=search)
        )

    if selected_status == "active":

        parents = parents.filter(
            is_active=True
        )

    elif selected_status == "inactive":

        parents = parents.filter(
            is_active=False
        )

    total_parents = Parent.objects.count()

    active_parents = Parent.objects.filter(
        is_active=True
    ).count()

    inactive_parents = Parent.objects.filter(
        is_active=False
    ).count()

    context = {

        "parents": parents,

        "search": search,

        "selected_status": selected_status,

        "total_parents": total_parents,

        "active_parents": active_parents,

        "inactive_parents": inactive_parents,

    }

    return render(
        request,
        "parents/parent_management.html",
        context
    )


# ============================================================
# FEE STRUCTURES
# ============================================================

@staff_member_required
def fee_structure_list(request):

    fee_structures = (
        FeeStructure.objects
        .select_related(
            "session",
            "term",
        )
        .order_by(
            "student_class",
            "student_type",
            "department",
        )
    )

    search = request.GET.get(
        "search",
        ""
    ).strip()

    selected_session = request.GET.get(
        "session",
        ""
    )

    selected_term = request.GET.get(
        "term",
        ""
    )

    selected_type = request.GET.get(
        "student_type",
        ""
    )

    if search:

        fee_structures = fee_structures.filter(
            Q(student_class__icontains=search)
            |
            Q(department__icontains=search)
            |
            Q(session__name__icontains=search)
            |
            Q(term__name__icontains=search)
        )

    if selected_session:

        fee_structures = fee_structures.filter(
            session_id=selected_session
        )

    if selected_term:

        fee_structures = fee_structures.filter(
            term_id=selected_term
        )

    if selected_type:

        fee_structures = fee_structures.filter(
            student_type=selected_type
        )

    context = {

        "fee_structures": fee_structures,

        "search": search,

        "selected_session": selected_session,

        "selected_term": selected_term,

        "selected_type": selected_type,

        "sessions": (
            AcademicSession.objects
            .all()
            .order_by(
                "-is_active",
                "-created_at"
            )
        ),

        "terms": (
            Term.objects
            .all()
            .order_by("id")
        ),

        "student_type_choices": (
            FeeStructure.STUDENT_TYPE_CHOICES
        ),

        "total_fee_structures": (
            FeeStructure.objects.count()
        ),

        "active_session": (
            AcademicSession.objects
            .filter(is_active=True)
            .first()
        ),
    }

    return render(
        request,
        "fees/fee_structure_management.html",
        context
    )


# ============================================================
# PAYMENTS
# ============================================================

@staff_member_required
def payment_list(request):

    payments = (
        Payment.objects
        .select_related(
            "student",
            "fee_structure",
        )
        .order_by(
            "-transaction_date"
        )
    )

    return render(
        request,
        "dashboard/payments.html",
        {
            "payments": payments,
        }
    )


# ============================================================
# PAYMENT DETAIL
# ============================================================

@staff_member_required
def payment_detail(request, payment_id):

    payment = get_object_or_404(
        Payment.objects.select_related(
            "student",
            "fee_structure",
        ),
        id=payment_id
    )

    return render(
        request,
        "dashboard/payment_detail.html",
        {
            "payment": payment,
        }
    )


# ============================================================
# DOWNLOAD PAYMENTS AS EXCEL
# ============================================================

@staff_member_required
def download_payments_excel(request):

    payments = (
        Payment.objects
        .select_related(
            "student",
            "fee_structure",
        )
        .order_by(
            "-transaction_date"
        )
    )

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Payments"

    worksheet.merge_cells("A1:H1")

    worksheet["A1"] = (
        "ROCK FEE PORTAL - PAYMENT REPORT"
    )

    worksheet["A1"].font = Font(
        bold=True,
        size=16
    )

    worksheet["A1"].alignment = Alignment(
        horizontal="center"
    )

    headers = [
        "Payment Reference",
        "Student Name",
        "Student ID",
        "Class",
        "Payment Type",
        "Amount",
        "Status",
        "Transaction Date",
    ]

    for column_number, header in enumerate(
        headers,
        start=1
    ):

        cell = worksheet.cell(
            row=3,
            column=column_number
        )

        cell.value = header

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    row_number = 4

    for payment in payments:

        student = payment.student

        worksheet.cell(
            row=row_number,
            column=1
        ).value = payment.payment_reference

        worksheet.cell(
            row=row_number,
            column=2
        ).value = student.full_name

        worksheet.cell(
            row=row_number,
            column=3
        ).value = student.student_id

        worksheet.cell(
            row=row_number,
            column=4
        ).value = student.student_class

        worksheet.cell(
            row=row_number,
            column=5
        ).value = payment.get_payment_type_display()

        worksheet.cell(
            row=row_number,
            column=6
        ).value = float(payment.amount)

        worksheet.cell(
            row=row_number,
            column=7
        ).value = payment.status

        transaction_date = payment.transaction_date

        if transaction_date:

            if timezone.is_aware(transaction_date):

                transaction_date = timezone.localtime(
                    transaction_date
                ).replace(
                    tzinfo=None
                )

        worksheet.cell(
            row=row_number,
            column=8
        ).value = transaction_date

        row_number += 1

    for row in worksheet.iter_rows(
        min_row=4,
        max_row=worksheet.max_row
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="center"
            )

    for row in range(
        4,
        worksheet.max_row + 1
    ):

        worksheet.cell(
            row=row,
            column=6
        ).number_format = "₦#,##0.00"

        worksheet.cell(
            row=row,
            column=8
        ).number_format = (
            "dd mmm yyyy hh:mm AM/PM"
        )

    column_widths = {
        "A": 25,
        "B": 28,
        "C": 18,
        "D": 15,
        "E": 22,
        "F": 18,
        "G": 15,
        "H": 25,
    }

    for column, width in column_widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    worksheet.freeze_panes = "A4"

    worksheet.auto_filter.ref = (
        f"A3:H{worksheet.max_row}"
    )

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; '
        'filename="rock_fee_payments.xlsx"'
    )

    workbook.save(response)

    return response


# ============================================================
# DOWNLOAD STUDENTS AS EXCEL
# ============================================================

@staff_member_required
def download_students_excel(request):

    students = (
        Student.objects
        .all()
        .order_by(
            "student_class",
            "full_name"
        )
    )

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Students"

    worksheet.merge_cells("A1:J1")

    worksheet["A1"] = (
        "ROCK FOUNDATION - STUDENT REPORT"
    )

    worksheet["A1"].font = Font(
        bold=True,
        size=16
    )

    worksheet["A1"].alignment = Alignment(
        horizontal="center"
    )

    headers = [
        "Student ID",
        "Full Name",
        "Class",
        "Student Type",
        "Department",
        "Parent / Guardian",
        "Parent Phone",
        "Parent Email",
        "Status",
        "Registration Date",
    ]

    for column_number, header in enumerate(
        headers,
        start=1
    ):

        cell = worksheet.cell(
            row=3,
            column=column_number
        )

        cell.value = header

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    row_number = 4

    for student in students:

        worksheet.cell(
            row=row_number,
            column=1
        ).value = student.student_id

        worksheet.cell(
            row=row_number,
            column=2
        ).value = student.full_name

        worksheet.cell(
            row=row_number,
            column=3
        ).value = student.student_class

        worksheet.cell(
            row=row_number,
            column=4
        ).value = student.get_student_type_display()

        worksheet.cell(
            row=row_number,
            column=5
        ).value = (
            student.get_department_display()
            if student.department
            else "Not Applicable"
        )

        worksheet.cell(
            row=row_number,
            column=6
        ).value = student.parent_name

        worksheet.cell(
            row=row_number,
            column=7
        ).value = student.parent_phone

        worksheet.cell(
            row=row_number,
            column=8
        ).value = student.parent_email

        worksheet.cell(
            row=row_number,
            column=9
        ).value = (
            "ACTIVE"
            if student.is_active
            else "INACTIVE"
        )

        registration_date = student.registered_at

        if registration_date:

            if timezone.is_aware(registration_date):

                registration_date = timezone.localtime(
                    registration_date
                ).replace(
                    tzinfo=None
                )

        worksheet.cell(
            row=row_number,
            column=10
        ).value = registration_date

        row_number += 1

    for row in worksheet.iter_rows(
        min_row=4,
        max_row=worksheet.max_row
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="center"
            )

    for row in range(
        4,
        worksheet.max_row + 1
    ):

        worksheet.cell(
            row=row,
            column=10
        ).number_format = (
            "dd mmm yyyy hh:mm AM/PM"
        )

    column_widths = {
        "A": 18,
        "B": 30,
        "C": 15,
        "D": 18,
        "E": 20,
        "F": 28,
        "G": 20,
        "H": 35,
        "I": 15,
        "J": 25,
    }

    for column, width in column_widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    worksheet.freeze_panes = "A4"

    worksheet.auto_filter.ref = (
        f"A3:J{worksheet.max_row}"
    )

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; '
        'filename="rock_foundation_students.xlsx"'
    )

    workbook.save(response)

    return response


# ============================================================
# DOWNLOAD PARENTS AS EXCEL
# ============================================================

@staff_member_required
def download_parents_excel(request):

    parents = (
        Parent.objects
        .select_related("user")
        .prefetch_related("students")
        .order_by("full_name")
    )

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Parents"

    worksheet.merge_cells("A1:H1")

    worksheet["A1"] = (
        "ROCK FOUNDATION - PARENT REPORT"
    )

    worksheet["A1"].font = Font(
        bold=True,
        size=16
    )

    worksheet["A1"].alignment = Alignment(
        horizontal="center"
    )

    headers = [
        "Parent Name",
        "Phone",
        "Email",
        "Username",
        "Students",
        "Number of Students",
        "Status",
        "Registration Date",
    ]

    for column_number, header in enumerate(
        headers,
        start=1
    ):

        cell = worksheet.cell(
            row=3,
            column=column_number
        )

        cell.value = header

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    row_number = 4

    for parent in parents:

        students = list(
            parent.students.all()
        )

        student_names = ", ".join(
            student.full_name
            for student in students
        )

        worksheet.cell(
            row=row_number,
            column=1
        ).value = parent.full_name

        worksheet.cell(
            row=row_number,
            column=2
        ).value = parent.phone

        worksheet.cell(
            row=row_number,
            column=3
        ).value = parent.user.email

        worksheet.cell(
            row=row_number,
            column=4
        ).value = parent.user.username

        worksheet.cell(
            row=row_number,
            column=5
        ).value = student_names or "None"

        worksheet.cell(
            row=row_number,
            column=6
        ).value = len(students)

        worksheet.cell(
            row=row_number,
            column=7
        ).value = (
            "ACTIVE"
            if parent.is_active
            else "INACTIVE"
        )

        created_date = parent.created_at

        if created_date:

            if timezone.is_aware(created_date):

                created_date = timezone.localtime(
                    created_date
                ).replace(
                    tzinfo=None
                )

        worksheet.cell(
            row=row_number,
            column=8
        ).value = created_date

        row_number += 1

    for row in worksheet.iter_rows(
        min_row=4,
        max_row=worksheet.max_row
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

    for row in range(
        4,
        worksheet.max_row + 1
    ):

        worksheet.cell(
            row=row,
            column=8
        ).number_format = (
            "dd mmm yyyy hh:mm AM/PM"
        )

    column_widths = {
        "A": 28,
        "B": 20,
        "C": 35,
        "D": 25,
        "E": 45,
        "F": 20,
        "G": 15,
        "H": 25,
    }

    for column, width in column_widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    worksheet.freeze_panes = "A4"

    worksheet.auto_filter.ref = (
        f"A3:H{worksheet.max_row}"
    )

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; '
        'filename="rock_foundation_parents.xlsx"'
    )

    workbook.save(response)

    return response